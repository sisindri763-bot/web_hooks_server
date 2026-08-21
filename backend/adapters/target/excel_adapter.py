"""
adapters/target/excel_adapter.py
----------------------------------
Excel target adapter — role="TARGET".
"""

import datetime
import io
import logging
import os
from typing import Any, Dict, Optional

import requests
from tenacity import retry, stop_after_attempt, wait_exponential, before_sleep_log

try:
    import openpyxl  # type: ignore # pyright: ignore[reportMissingImports]
except ImportError:
    openpyxl = None

from adapters.target.base import DataAdapter

logger = logging.getLogger(__name__)


class ExcelTargetAdapter(DataAdapter):
    role = "target"

    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=1, min=2, max=10),
        before_sleep=before_sleep_log(logger, logging.WARNING),
        reraise=True,
    )
    def fetch_snapshot(self, config: Dict[str, Any], run_id: Optional[str] = None) -> Dict[str, Any]:
        url_or_path = config["url_or_path"]
        sheet_name  = config.get("sheet_name")
        sample_n    = int(config.get("sample_rows", 10))
        object_name = url_or_path.rstrip("/").split("/")[-1]

        if openpyxl is None:
            return {
                "run_id":          run_id,
                "asset_role":      "TARGET",
                "system_name":     "Excel",
                "system_type":     "FILE",
                "database_name":   None,
                "schema_name":     sheet_name,
                "object_name":     object_name,
                "object_type":     "FILE",
                "row_count":       -1,
                "column_count":    -1,
                "size_bytes":      None,
                "last_updated_at": None,
                "observed_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
                "columns":         [],
                "sample":          [],
                "error":           "openpyxl not installed",
            }

        if url_or_path.startswith("http://") or url_or_path.startswith("https://"):
            resp       = requests.get(url_or_path, timeout=30)
            resp.raise_for_status()
            size_bytes = int(resp.headers.get("Content-Length", len(resp.content)))
            file_bytes = resp.content
        else:
            size_bytes = os.path.getsize(url_or_path)
            with open(url_or_path, "rb") as f:
                file_bytes = f.read()

        wb  = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        if ws is None:
            ws = wb.worksheets[0]
        sheet_used = getattr(ws, "title", "Sheet1")

        rows_iter = ws.iter_rows(values_only=True)  # type: ignore
        header    = [str(c) if c is not None else "" for c in next(rows_iter, [])]
        all_data  = [dict(zip(header, row)) for row in rows_iter]
        wb.close()

        return {
            "run_id":          run_id,
            "asset_role":      "TARGET",
            "system_name":     "Excel",
            "system_type":     "FILE",
            "database_name":   None,
            "schema_name":     sheet_used,
            "object_name":     object_name,
            "object_type":     "FILE",
            "row_count":       len(all_data),
            "column_count":    len(header),
            "size_bytes":      size_bytes,
            "last_updated_at": None,
            "observed_at":     datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "columns":         header,
            "sample":          all_data[:sample_n],
        }
